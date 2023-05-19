'''open('路径','操作文件的模式')
file.write()
file.read()
操作流程：1.打开文件--2.读或者写--3.关闭文件
ValueError: I/O operation on closed file.说明文件是关闭状态
*****: w模式会覆盖内容，a模式在原有内容新增
      +的区别就是即可读又可以写入，注意鼠标的指针位置
      write:
      read:
      a
      b:代表二进制
'''

'''文件操作w模式：覆盖写入,写入后鼠标指针在最后
w+:写+读（）
'''
# file=open(file='./test.txt',mode='w+',encoding='utf-8')  #file代表的是一个文件对象：txt，excel
# file.write('456')
# file.seek(0)
# print(file.tell())
# print('读取所有的内容：',file.read())
# file.close()


'''文件操作a模式:追加写入，不会覆盖文件
a+:读+写'''
# file=open(file='./test.txt',mode='a+',encoding='utf-8')  #file代表的是一个文件对象：txt，excel
# # file.write('\n第二行的数据')
# # file.write('\n第三行的数据')
# file.write('hjh')
# tell=file.tell()#获取文件指针位置
# print(tell)
# file.seek(0)
# print('读取所有的内容：',file.read())
# file.close()


'''文件操作r模式：读取 ,鼠标指针默认在最前面
文件路径一定要准确
r+:读+写
'''
# file=open(file='./test.txt',mode='r+',encoding='utf-8')  #file代表的是一个文件对象：txt，excel
# res1=file.read() #读取所有的内容，参数n表示指定长度
# print('文件所有的内容：',res1,type(res1))
# print(file.tell())
# file.write('111')
# print(file.tell())
#res2=file.readline() #读取一行内容，
#print('按行读取：',res2)
# res3=file.readlines()
# print('读取所有的行:',res3)
# for line in file.readlines():
#     print(line)


'''理解seek和tell方法'''
# tell=file.tell()#获取文件指针位置
# print('读取文件之前指针的位置：',tell)
#
# res1=file.read() #读取所有的内容，参数n表示指定长度
# print('文件所有的内容：',res1,type(res1))
# tell=file.tell()#获取文件指针位置
# print('读取文件之后指针的位置：',tell)
#
#
# file.seek(3)#移动文件读取指针到制定位置
# tell=file.tell()#获取文件指针位置
# print('调用了seek方法之后指针的位置：',tell)
#
# print(file.read())


'''with'''
# with open('test.txt',mode='r') as file:
#     print(file.readlines())


'''xlwt:excel表格的写入,不支持读取，覆盖写入，而不是追加
报错： [Errno 13] Permission denied: './test.xlsx'说明你的文件打开的状态，先保存关闭文件
'''
# import xlwt
# book=xlwt.Workbook(encoding='utf-8') #1.新建一个excel对象
# sheet_test=book.add_sheet('test1') #2.添加一个sheet工作表,表格名称叫sheet_test
# # sheet_test.write(0,0,'username')#3.给每个单元格添加值，单元格行和列分别从0开始(0,0)表示第一行第一列的单元格
# # sheet_test.write(0,1,'password')#4.给每个单元格添加值，单元格行和列分别从0开始(0,0)表示第一行第一列的单元格
# # sheet_test.write(0,2,'phone')#3.给每个单元格添加值，单元格行和列分别从0开始(0,0)表示第一行第一列的单元格
#
# '''写入一行的数据，把数据放到一个列表'''
# l=['username','password','phone']
# for i in range(0,len(l)):
#     print(i,l[i])
#     sheet_test.write(1, i, l[i])
# book.save('./test.xlsx')#4.保存文件： book.save(文件名.xls)




'''xlrd:只支持读取，不支持写入,安装的时候指定版本1.2.0
'''
# import xlrd
# excel=xlrd.open_workbook(filename='./test.xlsx')#打开一个文件(.xls或者.xlsx文件)
# sheet_test=excel.sheet_by_name('test1')#使用名称为test1的sheet工作表
# sheet_test1=excel.sheet_by_index(0)#使用sheet工作表,0代表第一个表格
#print(sheet_test.name,sheet_test1.name)
#sheet_all=excel.sheet_names()#xls/xlsx文件所有sheet名称：返回的是一个列表
#print(sheet_all)
#当前sheet的总行数：sh1.nrows； 当前sheet总列数：sh1.ncols
#print(sheet_test.nrows,sheet_test.ncols)
#取某一行：row1=sh1.row_values(0)
#print(sheet_test.row_values(1,0,3))
#sh1.cell_value(0,0)
# print(sheet_test.cell_value(1,1))


'''openpyxl:即可读取也可以写入，而且不会覆盖原有的数据:
与xlrd，xlwt的区别
    只支持xlsx格式，支持excel函数公式
    必须操作有一个已经存在的excel文件
    索引从1开始

'''
import openpyxl
book = openpyxl.load_workbook('test.xlsx')#打开文件(.xlsx文件),book表格excel对象
# #sheet_test=book.active #book.active或者
# #sheet_test=book.get_sheet_by_name('test1')
sheet_test=book['test1']
# # print(sheet_test.title)
#
# #4.写入数据
# #sheet_test['F4'] = 'daomingsi'
# #sheet_test.cell(1,2).value='success'
#
# #写入一行数据
# new_row = ['post-xml接口', 'post', 'https://httpbin.org/post']
# sheet_test.append(new_row)
#
# #5.保存数据
# book.save(filename='test.xlsx')



'''openpyxl读取
1.导包：import openpyxl
2.打开文件(.xlsx文件)：book = openpyxl.load_workbook(文件名)
使用sheet工作表：sh1=book.active或者sh1=book.get_sheet_by_name('Sheet1')  或者sh1=book['Sheet1']
读取sheet工作表的属性信息
当前sheet的名称：sh1.title
当前sheet的总行数：sh1.max_row
当前sheet的总列数：sh1.max_column
xlsx文件所有sheet名称：book.sheetnames
'''
# print(book.sheetnames)
# print(sheet_test.title)
# print(sheet_test.max_row)
# print(sheet_test.max_column)

#读取单元格数据
# cell1 = sheet_test['F4'].value
# print(cell1)
#print(sheet_test.cell(4,6).value)


'''
1.homework:写入一列数据['a','b','c',]

2.通过xlrd读取excel文件
usernamer	password	phone
admin	123456	1638749385094

写一个函数，返回的值

[
    {
    "username":"admin",
    "password":"123456",
    "phone":"1638749385094",
     },   
]

'''